"""
build_aluminum_library.py
Generates aluminum_expanded.xlsx — the energy reference library.

Run this OUTSIDE of NX (plain Python) to create/recreate the library.
Then copy aluminum_expanded.xlsx to your tool directory.

Data is empirical placeholder values — replace with your IN-MaC measurements.
Extended ranges cover: holes 2-75mm dia, chamfers to 50mm, fillets to 30mm,
taps up to M64, pockets to 10000 mm² area.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

OUTPUT_FILE = "aluminum_expanded.xlsx"


def make_header(ws, headers, color="4472C4"):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")


def write_rows(ws, rows, start_row=2):
    for r, row_data in enumerate(rows, start_row):
        for c, val in enumerate(row_data, 1):
            ws.cell(r, c, val)


# ============================================================================
def build_metadata(wb):
    ws = wb.create_sheet("metadata")
    data = [
        ("library_name",    "Aluminum (6061-T6) Machining Energy Library"),
        ("version",         "2.0-expanded"),
        ("material",        "aluminum"),
        ("material_spec",   "6061-T6"),
        ("units_energy",    "Wh"),
        ("units_length",    "mm"),
        ("units_area",      "mm2"),
        ("units_volume",    "mm3"),
        ("source",          "IN-MaC CNC monitoring + placeholder"),
        ("notes",           "Replace placeholder values with IN-MaC UUID measurements"),
    ]
    ws["A1"] = "key"
    ws["B1"] = "value"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for r, (k, v) in enumerate(data, 2):
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)


# ============================================================================
def build_thread_mapping(wb):
    ws = wb.create_sheet("thread_mapping")
    make_header(ws, ["thread_size_str", "diameter_mm", "notes"], "70AD47")
    rows = [
        # Metric
        ("M2",   2.0,   "Metric"),
        ("M2.5", 2.5,   "Metric"),
        ("M3",   3.0,   "Metric"),
        ("M4",   4.0,   "Metric"),
        ("M5",   5.0,   "Metric"),
        ("M6",   6.0,   "Metric"),
        ("M8",   8.0,   "Metric"),
        ("M10",  10.0,  "Metric"),
        ("M12",  12.0,  "Metric"),
        ("M14",  14.0,  "Metric"),
        ("M16",  16.0,  "Metric"),
        ("M18",  18.0,  "Metric"),
        ("M20",  20.0,  "Metric"),
        ("M22",  22.0,  "Metric"),
        ("M24",  24.0,  "Metric"),
        ("M27",  27.0,  "Metric"),
        ("M30",  30.0,  "Metric"),
        ("M33",  33.0,  "Metric"),
        ("M36",  36.0,  "Metric"),
        ("M39",  39.0,  "Metric"),
        ("M42",  42.0,  "Metric"),
        ("M48",  48.0,  "Metric"),
        ("M52",  52.0,  "Metric"),
        ("M56",  56.0,  "Metric"),
        ("M64",  64.0,  "Metric"),
        # Imperial UNC (common)
        ("#4-40",   2.845, "UNC"),
        ("#6-32",   3.505, "UNC"),
        ("#8-32",   4.166, "UNC"),
        ("#10-32",  4.826, "UNC"),
        ("1/4-20",  6.35,  "UNC"),
        ("5/16-18", 7.938, "UNC"),
        ("3/8-16",  9.525, "UNC"),
        ("7/16-14", 11.113,"UNC"),
        ("1/2-13",  12.7,  "UNC"),
        ("9/16-12", 14.288,"UNC"),
        ("5/8-11",  15.875,"UNC"),
        ("3/4-10",  19.05, "UNC"),
        ("7/8-9",   22.225,"UNC"),
        ("1-8",     25.4,  "UNC"),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_hole_simple(wb):
    """2D: diameter_mm × depth_mm → energy_wh"""
    ws = wb.create_sheet("hole_simple")
    make_header(ws, ["diameter_mm", "depth_mm", "energy_wh", "source", "uuid", "notes"])
    # (dia, depth, energy_wh)
    rows = [
        (2,   5,    0.010, "placeholder", "", ""),
        (2,   15,   0.020, "placeholder", "", ""),
        (3,   5,    0.012, "placeholder", "", ""),
        (3,   20,   0.025, "placeholder", "", ""),
        (4,   10,   0.018, "placeholder", "", ""),
        (4,   30,   0.038, "placeholder", "", ""),
        (5,   10,   0.022, "placeholder", "", ""),
        (5,   30,   0.045, "placeholder", "", ""),
        (6,   10,   0.028, "placeholder", "", ""),
        (6,   30,   0.058, "placeholder", "", ""),
        (6,   50,   0.090, "placeholder", "", ""),
        (8,   10,   0.038, "placeholder", "", ""),
        (8,   30,   0.072, "placeholder", "", ""),
        (8,   50,   0.110, "placeholder", "", ""),
        (10,  10,   0.048, "placeholder", "", ""),
        (10,  30,   0.090, "placeholder", "", ""),
        (10,  50,   0.140, "placeholder", "", ""),
        (10,  75,   0.200, "placeholder", "", ""),
        (12,  15,   0.065, "placeholder", "", ""),
        (12,  40,   0.120, "placeholder", "", ""),
        (12,  75,   0.220, "placeholder", "", ""),
        (15,  15,   0.085, "placeholder", "", ""),
        (15,  40,   0.155, "placeholder", "", ""),
        (15,  75,   0.275, "placeholder", "", ""),
        (18,  20,   0.110, "placeholder", "", ""),
        (18,  50,   0.200, "placeholder", "", ""),
        (18,  100,  0.370, "placeholder", "", ""),
        (20,  20,   0.130, "placeholder", "", ""),
        (20,  50,   0.230, "placeholder", "", ""),
        (20,  100,  0.420, "placeholder", "", ""),
        (25,  25,   0.180, "placeholder", "", ""),
        (25,  50,   0.310, "placeholder", "", ""),
        (25,  100,  0.560, "placeholder", "", ""),
        (30,  25,   0.220, "placeholder", "", ""),
        (30,  50,   0.385, "placeholder", "", ""),
        (30,  100,  0.700, "placeholder", "", ""),
        (35,  30,   0.280, "placeholder", "", ""),
        (35,  75,   0.540, "placeholder", "", ""),
        (40,  30,   0.340, "placeholder", "", ""),
        (40,  75,   0.640, "placeholder", "", ""),
        (50,  30,   0.460, "placeholder", "", ""),
        (50,  75,   0.860, "placeholder", "", ""),
        (60,  30,   0.600, "placeholder", "", ""),
        (60,  75,   1.100, "placeholder", "", ""),
        (75,  30,   0.800, "placeholder", "", ""),
        (75,  75,   1.500, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_counterbore_operation(wb):
    """2D: cb_diameter_mm × cb_depth_mm → energy_wh"""
    ws = wb.create_sheet("counterbore_operation")
    make_header(ws, ["cb_diameter_mm", "cb_depth_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (6,  3,  0.025, "placeholder", "", ""),
        (6,  8,  0.048, "placeholder", "", ""),
        (8,  3,  0.030, "placeholder", "", ""),
        (8,  8,  0.058, "placeholder", "", ""),
        (10, 4,  0.040, "placeholder", "", ""),
        (10, 10, 0.078, "placeholder", "", ""),
        (13, 5,  0.055, "placeholder", "", ""),
        (13, 12, 0.105, "placeholder", "", ""),
        (16, 6,  0.072, "placeholder", "", ""),
        (16, 14, 0.138, "placeholder", "", ""),
        (20, 8,  0.100, "placeholder", "", ""),
        (20, 18, 0.188, "placeholder", "", ""),
        (25, 10, 0.135, "placeholder", "", ""),
        (25, 22, 0.250, "placeholder", "", ""),
        (30, 10, 0.165, "placeholder", "", ""),
        (30, 25, 0.305, "placeholder", "", ""),
        (40, 12, 0.230, "placeholder", "", ""),
        (40, 30, 0.420, "placeholder", "", ""),
        (50, 15, 0.300, "placeholder", "", ""),
        (50, 35, 0.560, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_countersink_operation(wb):
    """1D: cs_diameter_mm → energy_wh"""
    ws = wb.create_sheet("countersink_operation")
    make_header(ws, ["cs_diameter_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (4,   0.008, "placeholder", "", "90° countersink"),
        (6,   0.012, "placeholder", "", ""),
        (8,   0.018, "placeholder", "", ""),
        (10,  0.025, "placeholder", "", ""),
        (13,  0.035, "placeholder", "", ""),
        (16,  0.048, "placeholder", "", ""),
        (20,  0.065, "placeholder", "", ""),
        (25,  0.085, "placeholder", "", ""),
        (30,  0.110, "placeholder", "", ""),
        (40,  0.155, "placeholder", "", ""),
        (50,  0.210, "placeholder", "", ""),
        (60,  0.270, "placeholder", "", ""),
        (75,  0.360, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_tap_operation(wb):
    """2D: thread_size_mm (numeric dia) × thread_depth_mm → energy_wh"""
    ws = wb.create_sheet("tap_operation")
    make_header(ws, ["thread_size_mm", "thread_depth_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (2,   5,   0.015, "placeholder", "", "M2"),
        (2,   15,  0.030, "placeholder", "", ""),
        (3,   5,   0.018, "placeholder", "", "M3"),
        (3,   20,  0.038, "placeholder", "", ""),
        (4,   8,   0.025, "placeholder", "", "M4"),
        (4,   25,  0.055, "placeholder", "", ""),
        (5,   8,   0.032, "placeholder", "", "M5"),
        (5,   25,  0.068, "placeholder", "", ""),
        (6,   10,  0.042, "placeholder", "", "M6"),
        (6,   30,  0.090, "placeholder", "", ""),
        (8,   12,  0.062, "placeholder", "", "M8"),
        (8,   35,  0.135, "placeholder", "", ""),
        (10,  15,  0.085, "placeholder", "", "M10"),
        (10,  40,  0.185, "placeholder", "", ""),
        (12,  18,  0.115, "placeholder", "", "M12"),
        (12,  50,  0.250, "placeholder", "", ""),
        (14,  20,  0.140, "placeholder", "", "M14"),
        (14,  55,  0.305, "placeholder", "", ""),
        (16,  22,  0.170, "placeholder", "", "M16"),
        (16,  60,  0.370, "placeholder", "", ""),
        (18,  25,  0.205, "placeholder", "", "M18"),
        (18,  65,  0.445, "placeholder", "", ""),
        (20,  28,  0.245, "placeholder", "", "M20"),
        (20,  70,  0.530, "placeholder", "", ""),
        (22,  30,  0.290, "placeholder", "", "M22"),
        (22,  75,  0.625, "placeholder", "", ""),
        (24,  32,  0.335, "placeholder", "", "M24"),
        (24,  80,  0.730, "placeholder", "", ""),
        (27,  35,  0.400, "placeholder", "", "M27"),
        (30,  40,  0.475, "placeholder", "", "M30"),
        (36,  45,  0.640, "placeholder", "", "M36"),
        (42,  50,  0.840, "placeholder", "", "M42"),
        (48,  55,  1.080, "placeholder", "", "M48"),
        (56,  60,  1.400, "placeholder", "", "M56"),
        (64,  65,  1.780, "placeholder", "", "M64"),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_pocket_rectangular(wb):
    """2D: area_mm2 × depth_mm → energy_wh"""
    ws = wb.create_sheet("pocket_rectangular")
    make_header(ws, ["area_mm2", "depth_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (100,   5,   0.045, "placeholder", "", "10×10"),
        (100,   15,  0.090, "placeholder", "", ""),
        (100,   25,  0.140, "placeholder", "", ""),
        (400,   5,   0.090, "placeholder", "", "20×20"),
        (400,   15,  0.175, "placeholder", "", ""),
        (400,   25,  0.270, "placeholder", "", ""),
        (900,   5,   0.160, "placeholder", "", "30×30"),
        (900,   15,  0.310, "placeholder", "", ""),
        (900,   25,  0.480, "placeholder", "", ""),
        (1600,  5,   0.260, "placeholder", "", "40×40"),
        (1600,  15,  0.500, "placeholder", "", ""),
        (1600,  25,  0.780, "placeholder", "", ""),
        (2500,  5,   0.380, "placeholder", "", "50×50"),
        (2500,  15,  0.730, "placeholder", "", ""),
        (2500,  30,  1.100, "placeholder", "", ""),
        (4000,  5,   0.560, "placeholder", "", "~63×63"),
        (4000,  15,  1.080, "placeholder", "", ""),
        (4000,  30,  1.620, "placeholder", "", ""),
        (6400,  5,   0.820, "placeholder", "", "80×80"),
        (6400,  15,  1.580, "placeholder", "", ""),
        (6400,  30,  2.360, "placeholder", "", ""),
        (10000, 5,   1.200, "placeholder", "", "100×100"),
        (10000, 15,  2.300, "placeholder", "", ""),
        (10000, 30,  3.450, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_pocket_circular(wb):
    """2D: area_mm2 × depth_mm → energy_wh"""
    ws = wb.create_sheet("pocket_circular")
    make_header(ws, ["area_mm2", "depth_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (78,    5,   0.038, "placeholder", "", "d=10"),
        (78,    15,  0.075, "placeholder", "", ""),
        (314,   5,   0.080, "placeholder", "", "d=20"),
        (314,   15,  0.155, "placeholder", "", ""),
        (314,   30,  0.240, "placeholder", "", ""),
        (707,   5,   0.150, "placeholder", "", "d=30"),
        (707,   15,  0.290, "placeholder", "", ""),
        (707,   30,  0.450, "placeholder", "", ""),
        (1257,  5,   0.245, "placeholder", "", "d=40"),
        (1257,  15,  0.470, "placeholder", "", ""),
        (1963,  5,   0.360, "placeholder", "", "d=50"),
        (1963,  15,  0.695, "placeholder", "", ""),
        (1963,  30,  1.050, "placeholder", "", ""),
        (3848,  5,   0.640, "placeholder", "", "d=70"),
        (3848,  15,  1.240, "placeholder", "", ""),
        (5027,  5,   0.820, "placeholder", "", "d=80"),
        (5027,  15,  1.580, "placeholder", "", ""),
        (7854,  5,   1.200, "placeholder", "", "d=100"),
        (7854,  15,  2.300, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_slot(wb):
    """1D: volume_mm3 → energy_wh"""
    ws = wb.create_sheet("slot")
    make_header(ws, ["volume_mm3", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (100,   0.040, "placeholder", "", ""),
        (500,   0.110, "placeholder", "", ""),
        (1000,  0.190, "placeholder", "", ""),
        (3000,  0.420, "placeholder", "", ""),
        (5000,  0.640, "placeholder", "", ""),
        (10000, 1.100, "placeholder", "", ""),
        (20000, 1.900, "placeholder", "", ""),
        (50000, 4.200, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_groove(wb):
    """1D: volume_mm3 → energy_wh"""
    ws = wb.create_sheet("groove")
    make_header(ws, ["volume_mm3", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (50,    0.025, "placeholder", "", ""),
        (200,   0.065, "placeholder", "", ""),
        (500,   0.130, "placeholder", "", ""),
        (1000,  0.235, "placeholder", "", ""),
        (2000,  0.400, "placeholder", "", ""),
        (5000,  0.850, "placeholder", "", ""),
        (10000, 1.600, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_chamfer(wb):
    """1D (per edge): offset_mm → energy_wh"""
    ws = wb.create_sheet("chamfer")
    make_header(ws, ["offset_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (0.5,  0.005, "placeholder", "", "per edge"),
        (1.0,  0.010, "placeholder", "", ""),
        (2.0,  0.018, "placeholder", "", ""),
        (3.0,  0.028, "placeholder", "", ""),
        (5.0,  0.048, "placeholder", "", ""),
        (8.0,  0.078, "placeholder", "", ""),
        (10.0, 0.098, "placeholder", "", ""),
        (15.0, 0.148, "placeholder", "", ""),
        (20.0, 0.200, "placeholder", "", ""),
        (25.0, 0.258, "placeholder", "", ""),
        (30.0, 0.315, "placeholder", "", ""),
        (40.0, 0.430, "placeholder", "", ""),
        (50.0, 0.550, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_fillet(wb):
    """1D (per edge chain): radius_mm → energy_wh"""
    ws = wb.create_sheet("fillet")
    make_header(ws, ["radius_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (0.5,  0.003, "placeholder", "", "per chain"),
        (1.0,  0.005, "placeholder", "", ""),
        (2.0,  0.010, "placeholder", "", ""),
        (3.0,  0.015, "placeholder", "", ""),
        (5.0,  0.025, "placeholder", "", ""),
        (8.0,  0.042, "placeholder", "", ""),
        (10.0, 0.055, "placeholder", "", ""),
        (12.0, 0.068, "placeholder", "", ""),
        (15.0, 0.085, "placeholder", "", ""),
        (20.0, 0.115, "placeholder", "", ""),
        (25.0, 0.148, "placeholder", "", ""),
        (30.0, 0.185, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_face(wb):
    """2D: area_mm2 × depth_mm (doc passes) → energy_wh"""
    ws = wb.create_sheet("face")
    make_header(ws, ["area_mm2", "depth_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (1000,  0.5, 0.050, "placeholder", "", "1 pass"),
        (1000,  2.0, 0.110, "placeholder", "", "2 passes"),
        (5000,  0.5, 0.180, "placeholder", "", ""),
        (5000,  2.0, 0.380, "placeholder", "", ""),
        (10000, 0.5, 0.330, "placeholder", "", "100×100mm"),
        (10000, 2.0, 0.700, "placeholder", "", ""),
        (25000, 0.5, 0.750, "placeholder", "", ""),
        (25000, 2.0, 1.600, "placeholder", "", ""),
        (50000, 0.5, 1.400, "placeholder", "", ""),
        (50000, 2.0, 2.950, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def build_thread_external(wb):
    """1D: length_mm → energy_wh"""
    ws = wb.create_sheet("thread_external")
    make_header(ws, ["length_mm", "energy_wh", "source", "uuid", "notes"])
    rows = [
        (5,   0.012, "placeholder", "", ""),
        (10,  0.020, "placeholder", "", ""),
        (20,  0.038, "placeholder", "", ""),
        (30,  0.055, "placeholder", "", ""),
        (50,  0.090, "placeholder", "", ""),
        (75,  0.135, "placeholder", "", ""),
        (100, 0.180, "placeholder", "", ""),
    ]
    write_rows(ws, rows)


# ============================================================================
def main():
    wb = openpyxl.Workbook()
    # Remove default empty sheet
    default = wb.active
    wb.remove(default)

    print("Building sheets...")
    build_metadata(wb)
    build_thread_mapping(wb)
    build_hole_simple(wb)
    build_counterbore_operation(wb)
    build_countersink_operation(wb)
    build_tap_operation(wb)
    build_pocket_rectangular(wb)
    build_pocket_circular(wb)
    build_slot(wb)
    build_groove(wb)
    build_chamfer(wb)
    build_fillet(wb)
    build_face(wb)
    build_thread_external(wb)

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
