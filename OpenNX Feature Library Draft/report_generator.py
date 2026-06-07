"""
report_generator.py - Report Generation for CAD Energy Estimator

Formats estimation results into:
  - NX listing window (text)
  - JSON file
  - Excel workbook with per-feature breakdown

Version: 4.2 - Handles compound features, reads from energy_result dict,
               quality summary counts operations correctly.
"""

import os
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ReportGenerator:

    def __init__(self, listing_window=None, debug=False):
        self.lw    = listing_window
        self.debug = debug

    # =========================================================================
    # TEXT REPORT → listing window
    # =========================================================================

    def print_report(self, result: Dict):
        """Print full text report to NX listing window."""
        lines = self._format_result(result)
        if self.lw:
            for line in lines:
                try:
                    self.lw.WriteLine(line)
                except Exception:
                    pass

    def _format_result(self, result: Dict) -> List[str]:
        if result.get("type") == "assembly":
            return self._format_assembly(result)
        return self._format_part(result)

    def _format_assembly(self, result: Dict) -> List[str]:
        lines = []
        lines.append("")
        lines.append("=" * 70)
        lines.append(f"ASSEMBLY ENERGY REPORT: {result['name']}")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 70)
        lines.append("")

        total = 0.0
        for comp in result.get("components", []):
            lines += self._format_part(comp, indent="  ")
            total += comp.get("total_energy_wh", 0.0)

        skipped = result.get("purchased_parts_skipped", 0)
        if skipped:
            lines.append(f"  (Purchased parts skipped: {skipped})")

        lines.append("")
        lines.append("-" * 70)
        lines.append(f"ASSEMBLY TOTAL:  {total:.4f} Wh  ({total/1000:.6f} kWh)")
        lines.append("=" * 70)
        return lines

    def _format_part(self, result: Dict, indent: str = "") -> List[str]:
        lines = []
        name     = result.get("name", "UnnamedPart")
        total_wh = result.get("total_energy_wh", 0.0)

        if result.get("is_purchased"):
            lines.append(f"{indent}PART: {name}  [PURCHASED — skipped]")
            return lines

        lines.append("")
        lines.append(f"{indent}" + "=" * 60)
        lines.append(f"{indent}PART: {name}")
        lines.append(f"{indent}" + "-" * 60)

        features = result.get("features", [])
        for feat in features:
            lines += self._format_feature(feat, indent + "  ")

        lines.append(f"{indent}" + "-" * 60)
        lines.append(f"{indent}PART TOTAL:  {total_wh:.4f} Wh  ({total_wh/1000:.6f} kWh)")

        # Quality summary
        lines += self._format_quality_summary(features, indent)

        return lines

    def _format_feature(self, feat: Dict, indent: str = "") -> List[str]:
        lines = []
        jid      = feat.get("journal_id") or feat.get("name", "?")
        category = feat.get("category", "?")
        status   = feat.get("extraction_status", "ok")
        energy   = feat.get("energy_wh", 0.0)

        if feat.get("is_multiplier"):
            parent = feat.get("parent_name", "?")
            count  = feat.get("instance_count", "?")
            lines.append(f"{indent}{jid:<30} [{category}]  ×{count} copies of {parent}  "
                         f"= {energy:.4f} Wh")
            return lines

        lines.append(f"{indent}{jid:<30} [{category}]  {energy:.4f} Wh  [{status}]")

        if status == "error":
            lines.append(f"{indent}  ERROR: {feat.get('extraction_error','?')}")
            return lines

        energy_result = feat.get("energy_result", {})

        # Compound feature: show sub-operations indented
        ops = energy_result.get("operations", [])
        if ops:
            for op in ops:
                op_cat = op.get("category", "?")
                op_e   = op.get("energy_wh", 0.0)
                basis  = op.get("estimation_basis", "")
                lines.append(f"{indent}    {op_cat:<26} {op_e:.4f} Wh  [{basis}]")
                for w in op.get("warnings", []):
                    lines.append(f"{indent}    ! {w}")
        else:
            basis = energy_result.get("estimation_basis", "")
            if basis:
                lines.append(f"{indent}  Basis: {basis}")
            for w in energy_result.get("warnings", []):
                lines.append(f"{indent}  ! {w}")

        return lines

    # =========================================================================
    # QUALITY SUMMARY
    # =========================================================================

    def _format_quality_summary(self, features: List[Dict], indent: str = "") -> List[str]:
        lines = []
        if not features:
            return lines

        by_cat = defaultdict(lambda: {"count": 0, "energy": 0.0,
                                       "interp": 0, "extrap": 0, "miss": 0})

        def _tally(cat, energy_wh, basis, method, warnings):
            by_cat[cat]["count"]  += 1
            by_cat[cat]["energy"] += (energy_wh or 0.0)
            b = (basis or "").lower()
            m = (method or "").lower()
            if "extrapolat" in b or "extrapolat" in m:
                by_cat[cat]["extrap"] += 1
            elif "missing" in b or energy_wh == 0:
                by_cat[cat]["miss"]   += 1
            else:
                by_cat[cat]["interp"] += 1

        for f in features:
            if f.get("is_multiplier"):
                continue
            energy_result = f.get("energy_result", {})
            ops = energy_result.get("operations", [])
            if ops:
                for op in ops:
                    op_cat    = op.get("category", f.get("category", "unknown"))
                    op_e      = op.get("energy_wh", 0.0)
                    op_basis  = op.get("estimation_basis", "")
                    op_lookup = op.get("lookup_details") or {}
                    op_method = op_lookup.get("method", "")
                    _tally(op_cat, op_e, op_basis, op_method, op.get("warnings", []))
            else:
                cat    = f.get("category", "unknown")
                e      = energy_result.get("energy_wh", f.get("energy_wh", 0.0))
                basis  = energy_result.get("estimation_basis", "")
                lookup = energy_result.get("lookup_details") or {}
                method = lookup.get("method", "")
                _tally(cat, e, basis, method, energy_result.get("warnings", []))

        lines.append("")
        lines.append(f"{indent}QUALITY SUMMARY")
        lines.append(f"{indent}" + "-" * 64)
        lines.append(f"{indent}{'Category':<22} {'N':>4} {'Energy Wh':>10} "
                     f"{'Interp':>7} {'Extrap':>7} {'Miss':>5}")
        lines.append(f"{indent}" + "-" * 64)

        t_interp = t_extrap = t_miss = 0
        for cat, d in sorted(by_cat.items()):
            lines.append(f"{indent}{cat:<22} {d['count']:>4} {d['energy']:>10.4f} "
                         f"{d['interp']:>7} {d['extrap']:>7} {d['miss']:>5}")
            t_interp += d["interp"]
            t_extrap += d["extrap"]
            t_miss   += d["miss"]

        n_ops = t_interp + t_extrap + t_miss
        lines.append(f"{indent}" + "-" * 64)
        if n_ops > 0:
            lines.append(f"{indent}Coverage: {t_interp/n_ops*100:.0f}% interp, "
                         f"{t_extrap/n_ops*100:.0f}% extrap, "
                         f"{t_miss/n_ops*100:.0f}% missing")
        return lines

    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================

    def export_excel(self, result: Dict, filepath: str):
        """Export result to multi-sheet Excel workbook."""
        if not _HAS_OPENPYXL:
            if self.lw:
                self.lw.WriteLine("WARNING: openpyxl not available — Excel export skipped.")
            return

        wb = openpyxl.Workbook()
        ws_sum = wb.active
        ws_sum.title = "Summary"

        if result.get("type") == "assembly":
            self._excel_assembly(wb, ws_sum, result)
        else:
            self._excel_part_summary(ws_sum, result)
            self._excel_part_detail(wb, result)

        try:
            wb.save(filepath)
        except Exception as e:
            if self.lw:
                self.lw.WriteLine(f"Excel save error: {e}")

    def _excel_assembly(self, wb, ws_sum, result):
        """Write assembly summary sheet + one detail sheet per part."""
        self._style_header(ws_sum)
        ws_sum["A1"] = "Assembly Energy Report"
        ws_sum["A2"] = result["name"]
        ws_sum["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 5
        ws_sum.cell(row, 1, "Part Name")
        ws_sum.cell(row, 2, "Energy (Wh)")
        ws_sum.cell(row, 3, "Features")
        ws_sum.cell(row, 4, "Status")
        row += 1

        total = 0.0
        for comp in result.get("components", []):
            name     = comp.get("name", "?")
            energy   = comp.get("total_energy_wh", 0.0)
            n_feats  = len(comp.get("features", []))
            status   = "Purchased" if comp.get("is_purchased") else "Machined"
            ws_sum.cell(row, 1, name)
            ws_sum.cell(row, 2, round(energy, 4))
            ws_sum.cell(row, 3, n_feats)
            ws_sum.cell(row, 4, status)
            row += 1
            total += energy
            if not comp.get("is_purchased") and comp.get("features"):
                self._excel_part_detail(wb, comp)

        ws_sum.cell(row, 1, "TOTAL")
        ws_sum.cell(row, 2, round(total, 4))

    def _excel_part_summary(self, ws, result):
        """Summarise part on provided worksheet."""
        self._style_header(ws)
        ws["A1"] = "Part Energy Report"
        ws["A2"] = result.get("name", "")
        ws["A3"] = f"Total: {result.get('total_energy_wh', 0):.4f} Wh"
        ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    def _excel_part_detail(self, wb, result):
        """One sheet per part with row-per-feature detail."""
        sheet_name = (result.get("name") or "Part")[:31]
        # Ensure unique
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + "_1"

        ws = wb.create_sheet(title=sheet_name)

        headers = ["Feature ID", "Category", "Key Dim 1", "Value 1",
                   "Key Dim 2", "Value 2", "Energy (Wh)", "Basis", "Warnings"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for feat in result.get("features", []):
            rows = self._feature_to_excel_rows(feat)
            for r in rows:
                for col, val in enumerate(r, 1):
                    ws.cell(row, col, val)
                row += 1

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    def _feature_to_excel_rows(self, feat):
        """Convert feature dict to list of row tuples for Excel."""
        jid      = feat.get("journal_id") or feat.get("name", "?")
        category = feat.get("category", "?")
        energy   = feat.get("energy_wh", 0.0)
        er       = feat.get("energy_result", {})
        basis    = er.get("estimation_basis", "")
        warnings = "; ".join(er.get("warnings", []))

        ops = er.get("operations", [])
        if ops:
            rows = []
            rows.append([jid, category, "", "", "", "", round(energy, 6),
                         "Compound feature", ""])
            for op in ops:
                op_cat  = op.get("category", "?")
                op_e    = op.get("energy_wh", 0.0)
                op_b    = op.get("estimation_basis", "")
                op_w    = "; ".join(op.get("warnings", []))
                ld      = op.get("lookup_details") or {}
                dim1    = list(ld.get("dim1", {}).items())[0] if ld.get("dim1") else ("", "")
                dim2    = list(ld.get("dim2", {}).items())[0] if ld.get("dim2") else ("", "")
                rows.append(["  " + op_cat, op_cat,
                             dim1[0], _fmt(dim1[1]),
                             dim2[0], _fmt(dim2[1]),
                             round(op_e, 6), op_b, op_w])
            return rows

        ld   = er.get("lookup_details") or {}
        dim1 = list(ld.get("dim1", {}).items())[0] if ld.get("dim1") else ("", "")
        dim2 = list(ld.get("dim2", {}).items())[0] if ld.get("dim2") else ("", "")

        return [[jid, category,
                 dim1[0], _fmt(dim1[1]),
                 dim2[0], _fmt(dim2[1]),
                 round(energy, 6), basis, warnings]]

    # =========================================================================
    # STYLE HELPERS
    # =========================================================================

    def _style_header(self, ws):
        ws["A1"].font = Font(bold=True, size=14)


# ============================================================================
# UTILITY
# ============================================================================

def _fmt(val):
    if val is None:
        return ""
    try:
        return round(float(val), 4)
    except (TypeError, ValueError):
        return str(val)
