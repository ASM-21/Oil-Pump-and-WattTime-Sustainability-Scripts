"""
energy_lookup.py - Feature Energy Lookup with Interpolation

Loads empirical energy data from an Excel library and performs
1D or 2D interpolation to estimate energy for extracted features.

Version: 4.2 - compatible with feature_extractor v4.2 compound feature output
"""

import os
import re
import math
import openpyxl
from typing import Dict, List, Optional, Any


# ============================================================================
# FEATURE CONFIGURATION
# Maps internal category names to library sheets and interpolation parameters
# ============================================================================

FEATURE_CONFIG = {
    # --- Holes ---
    "hole_simple": {
        "sheet":     "hole_simple",
        "interp":    "2d",
        "dim1_key":  "diameter_mm",
        "dim2_key":  "depth_mm",
        "dim1_lib":  "diameter_mm",
        "dim2_lib":  "depth_mm",
    },
    "counterbore_operation": {
        "sheet":     "counterbore_operation",
        "interp":    "2d",
        "dim1_key":  "cb_diameter_mm",
        "dim2_key":  "cb_depth_mm",
        "dim1_lib":  "cb_diameter_mm",
        "dim2_lib":  "cb_depth_mm",
    },
    "countersink_operation": {
        "sheet":     "countersink_operation",
        "interp":    "1d",
        "dim1_key":  "cs_diameter_mm",
        "dim1_lib":  "cs_diameter_mm",
    },
    "tap_operation": {
        "sheet":     "tap_operation",
        "interp":    "2d_thread",
        "dim1_key":  "thread_size",
        "dim2_key":  "thread_depth_mm",
        "dim1_lib":  "thread_size_mm",
        "dim2_lib":  "thread_depth_mm",
    },
    # --- Pockets ---
    "pocket_rectangular": {
        "sheet":     "pocket_rectangular",
        "interp":    "2d",
        "dim1_key":  "area_mm2",
        "dim2_key":  "depth_mm",
        "dim1_lib":  "area_mm2",
        "dim2_lib":  "depth_mm",
    },
    "pocket_shallow": {
        "sheet":     "pocket_rectangular",
        "interp":    "2d",
        "dim1_key":  "area_mm2",
        "dim2_key":  "depth_mm",
        "dim1_lib":  "area_mm2",
        "dim2_lib":  "depth_mm",
    },
    "pocket_circular": {
        "sheet":     "pocket_circular",
        "interp":    "2d",
        "dim1_key":  "area_mm2",
        "dim2_key":  "depth_mm",
        "dim1_lib":  "area_mm2",
        "dim2_lib":  "depth_mm",
    },
    # --- Slot / Groove ---
    "slot": {
        "sheet":     "slot",
        "interp":    "1d",
        "dim1_key":  "volume_mm3",
        "dim1_lib":  "volume_mm3",
    },
    "groove": {
        "sheet":     "groove",
        "interp":    "1d",
        "dim1_key":  "volume_mm3",
        "dim1_lib":  "volume_mm3",
    },
    # --- Chamfer / Fillet ---
    "chamfer": {
        "sheet":     "chamfer",
        "interp":    "1d",
        "dim1_key":  "offset_mm",
        "dim1_lib":  "offset_mm",
        "per_edge":  True,
        "edge_key":  "edge_count",
    },
    "fillet": {
        "sheet":     "fillet",
        "interp":    "1d",
        "dim1_key":  "radius_mm",
        "dim1_lib":  "radius_mm",
        "per_edge":  True,
        "edge_key":  "edge_chain_count",
    },
    # --- Face / External Thread ---
    "face": {
        "sheet":     "face",
        "interp":    "2d",
        "dim1_key":  "area_mm2",
        "dim2_key":  "depth_mm",
        "dim1_lib":  "area_mm2",
        "dim2_lib":  "depth_mm",
    },
    "thread_external": {
        "sheet":     "thread_external",
        "interp":    "1d",
        "dim1_key":  "length_mm",
        "dim1_lib":  "length_mm",
    },
}


# ============================================================================
# LIBRARY LOADER
# ============================================================================

class FeatureLibraryLoader:
    """Load and expose energy reference data from an Excel file."""

    def __init__(self, excel_path: str):
        if not os.path.isfile(excel_path):
            raise FileNotFoundError(f"Library not found: {excel_path}")

        self.excel_path = excel_path
        self.workbook   = openpyxl.load_workbook(excel_path, data_only=True)
        self.metadata   = self._load_metadata()
        self.thread_map = self._load_thread_mapping()
        self.sheets     = {}

        for sheet_name in self.workbook.sheetnames:
            if sheet_name in ("metadata", "thread_mapping"):
                continue
            self.sheets[sheet_name] = self._load_sheet(sheet_name)

    def _load_metadata(self) -> Dict[str, str]:
        meta = {}
        if "metadata" in self.workbook.sheetnames:
            ws = self.workbook["metadata"]
            for row in ws.iter_rows(values_only=True):
                if row[0] and row[1]:
                    meta[str(row[0])] = str(row[1])
        return meta

    def _load_thread_mapping(self) -> Dict[str, float]:
        """thread_size_str → equivalent diameter in mm."""
        mapping = {}
        if "thread_mapping" in self.workbook.sheetnames:
            ws = self.workbook["thread_mapping"]
            header = True
            for row in ws.iter_rows(values_only=True):
                if header:
                    header = False
                    continue
                if row[0] and row[1]:
                    try:
                        mapping[str(row[0])] = float(row[1])
                    except (TypeError, ValueError):
                        pass
        return mapping

    def _load_sheet(self, sheet_name: str) -> List[Dict]:
        ws = self.workbook[sheet_name]
        rows = []
        headers = None

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [
                    str(h).strip().lower() if h is not None else f"col_{i}"
                    for i, h in enumerate(row)
                ]
                continue
            if row[0] is None:
                continue
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            rows.append(row_dict)

        return rows

    def get_points(self, sheet_name: str) -> List[Dict]:
        return self.sheets.get(sheet_name, [])


# ============================================================================
# ENERGY LOOKUP / INTERPOLATION ENGINE
# ============================================================================

class EnergyLookup:
    """Estimate energy for extracted features using empirical library."""

    def __init__(self, library: FeatureLibraryLoader):
        self.lib = library

    def estimate(self, feature: Dict) -> Dict:
        """
        Main entry: estimate energy for a feature dict.
        Handles simple features and compound (hole_compound) features.
        Returns an energy_result dict.
        """
        category = feature.get("category", "unknown")

        if category == "hole_compound":
            return self._estimate_compound(feature)

        cfg = FEATURE_CONFIG.get(category)
        if cfg is None:
            return {
                "energy_wh":         0.0,
                "estimation_basis":  f"Unknown category: {category}",
                "warnings":          [f"No library config for category '{category}'"],
            }

        return self._estimate_single(feature, cfg)

    # -------------------------------------------------------------------------
    # COMPOUND FEATURE (counterbore, countersink-threaded, etc.)
    # -------------------------------------------------------------------------

    def _estimate_compound(self, feature: Dict) -> Dict:
        ops = feature.get("decomposed_operations", [])
        total_energy = 0.0
        op_results = []
        warnings = []

        for op in ops:
            op_cat = op.get("category", "unknown")
            cfg    = FEATURE_CONFIG.get(op_cat)
            if cfg is None:
                op_result = {
                    "category":         op_cat,
                    "energy_wh":        0.0,
                    "estimation_basis": f"Unknown operation category: {op_cat}",
                    "warnings":         [f"No config for {op_cat}"],
                }
            else:
                op_result = self._estimate_single(op, cfg)
                op_result["category"] = op_cat

            total_energy += op_result.get("energy_wh", 0.0)
            warnings.extend(op_result.get("warnings", []))
            op_results.append(op_result)

        return {
            "energy_wh":        total_energy,
            "estimation_basis": "Compound feature: sum of constituent operations",
            "operations":       op_results,
            "warnings":         warnings,
        }

    # -------------------------------------------------------------------------
    # SINGLE FEATURE
    # -------------------------------------------------------------------------

    def _estimate_single(self, feature: Dict, cfg: Dict) -> Dict:
        interp_type = cfg.get("interp", "1d")

        if interp_type == "2d_thread":
            return self._estimate_thread(feature, cfg)
        elif interp_type == "2d":
            return self._estimate_2d(feature, cfg)
        else:
            return self._estimate_1d(feature, cfg)

    # -------------------------------------------------------------------------
    # 1D INTERPOLATION
    # -------------------------------------------------------------------------

    def _estimate_1d(self, feature: Dict, cfg: Dict) -> Dict:
        sheet_name = cfg["sheet"]
        dim1_key   = cfg["dim1_key"]
        dim1_lib   = cfg["dim1_lib"]

        dim1 = self._get_val(feature, dim1_key)
        if dim1 is None:
            return self._missing(f"Missing parameter: {dim1_key}")

        points = self.lib.get_points(sheet_name)
        if not points:
            return self._missing(f"Library sheet empty: {sheet_name}")

        # Sort reference points by dimension
        pts = []
        for p in points:
            x = self._num(p.get(dim1_lib))
            y = self._num(p.get("energy_wh"))
            if x is not None and y is not None:
                pts.append((x, y))
        pts.sort(key=lambda t: t[0])

        if not pts:
            return self._missing(f"No valid rows in {sheet_name}")

        energy, basis, extrap = self._interp1d(dim1, pts)

        # Per-edge multiplication
        multiplier = 1
        if cfg.get("per_edge"):
            edge_key = cfg.get("edge_key", "edge_count")
            multiplier = int(feature.get(edge_key) or 1)
            energy *= multiplier

        warnings = []
        if extrap:
            warnings.append(f"Extrapolated: {dim1_key}={dim1:.2f} outside range "
                            f"[{pts[0][0]:.2f}, {pts[-1][0]:.2f}]")

        return {
            "energy_wh":         energy,
            "estimation_basis":  basis,
            "extrap":            extrap,
            "edge_multiplier":   multiplier,
            "lookup_details": {
                "sheet":   sheet_name,
                "method":  "extrapolated" if extrap else "interpolated",
                "dim1":    {dim1_key: dim1},
            },
            "warnings": warnings,
        }

    # -------------------------------------------------------------------------
    # 2D INTERPOLATION (bilinear)
    # -------------------------------------------------------------------------

    def _estimate_2d(self, feature: Dict, cfg: Dict) -> Dict:
        sheet_name = cfg["sheet"]
        dim1_key   = cfg["dim1_key"]
        dim2_key   = cfg["dim2_key"]
        dim1_lib   = cfg["dim1_lib"]
        dim2_lib   = cfg["dim2_lib"]

        dim1 = self._get_val(feature, dim1_key)
        dim2 = self._get_val(feature, dim2_key)

        if dim1 is None:
            return self._missing(f"Missing parameter: {dim1_key}")

        points = self.lib.get_points(sheet_name)
        if not points:
            return self._missing(f"Library sheet empty: {sheet_name}")

        # Parse reference points
        pts = []
        for p in points:
            x = self._num(p.get(dim1_lib))
            y = self._num(p.get(dim2_lib))
            e = self._num(p.get("energy_wh"))
            if x is not None and e is not None:
                pts.append((x, y, e))

        if not pts:
            return self._missing(f"No valid rows in {sheet_name}")

        # If dim2 missing, fall back to 1D on dim1
        if dim2 is None:
            pts_1d = [(x, e) for x, _, e in pts]
            pts_1d.sort(key=lambda t: t[0])
            energy, basis, extrap = self._interp1d(dim1, pts_1d)
            basis += f" (1D fallback — {dim2_key} unknown)"
            warnings = [f"Missing {dim2_key}; used 1D interpolation on {dim1_key}"]
            if extrap:
                warnings.append(f"Extrapolated: {dim1_key}={dim1:.2f}")
            return {
                "energy_wh":        energy,
                "estimation_basis": basis,
                "extrap":           extrap,
                "lookup_details":   {"sheet": sheet_name, "method": "1d_fallback"},
                "warnings":         warnings,
            }

        energy, basis, extrap = self._interp2d(dim1, dim2, pts)
        warnings = []
        if extrap:
            warnings.append(f"Extrapolated: ({dim1_key}={dim1:.2f}, {dim2_key}={dim2:.2f})")

        return {
            "energy_wh":        energy,
            "estimation_basis": basis,
            "extrap":           extrap,
            "lookup_details": {
                "sheet":  sheet_name,
                "method": "extrapolated" if extrap else "interpolated",
                "dim1":   {dim1_key: dim1},
                "dim2":   {dim2_key: dim2},
            },
            "warnings": warnings,
        }

    # -------------------------------------------------------------------------
    # THREAD LOOKUP
    # -------------------------------------------------------------------------

    def _estimate_thread(self, feature: Dict, cfg: Dict) -> Dict:
        thread_size = feature.get("thread_size")
        if thread_size is None:
            return self._missing("Missing thread_size")

        dia_mm, parse_method = self._parse_thread_size(thread_size)
        if dia_mm is None:
            return self._missing(f"Cannot parse thread size: '{thread_size}'")

        # Inject numeric diameter into a temp dict for 2d lookup
        temp = dict(feature)
        temp[cfg["dim1_key"]] = dia_mm
        result = self._estimate_2d(temp, {
            **cfg,
            "interp":   "2d",
            "dim1_key": cfg["dim1_key"],
            "dim1_lib": cfg["dim1_lib"],
        })
        result.setdefault("lookup_details", {})["thread_parse_method"] = parse_method
        result.setdefault("lookup_details", {})["thread_size_str"] = thread_size
        result.setdefault("lookup_details", {})["thread_dia_mm"]   = dia_mm
        return result

    def _parse_thread_size(self, size_str: str):
        """
        Return (diameter_mm, method_name) for a thread size string.
        Tries: library mapping table, metric regex, imperial fraction, pipe.
        """
        s = str(size_str).strip()

        # 1. Mapping table (highest priority)
        if s in self.lib.thread_map:
            return self.lib.thread_map[s], "mapping_table"

        # 2. Metric: M6, M6x1.0, M6x1, m6
        m = re.match(r"[Mm](\d+(?:\.\d+)?)", s)
        if m:
            return float(m.group(1)), "metric_regex"

        # 3. Imperial unified: 1/4-20, 3/8-16, #10-32
        m = re.match(r"#?(\d+)/(\d+)-(\d+)", s)
        if m:
            num, den = int(m.group(1)), int(m.group(2))
            dia_in = num / den
            return dia_in * 25.4, "imperial_unified"

        # 4. Number sizes (#4, #6, #10)
        num_sizes = {"0": 1.524, "1": 1.854, "2": 2.184, "3": 2.515,
                     "4": 2.845, "5": 3.175, "6": 3.505, "8": 4.166,
                     "10": 4.826, "12": 5.486}
        m = re.match(r"#?(\d+)-\d+", s)
        if m and m.group(1) in num_sizes:
            return num_sizes[m.group(1)], "number_size"

        # 5. Decimal inch: 0.25-20
        m = re.match(r"(0\.\d+)-\d+", s)
        if m:
            return float(m.group(1)) * 25.4, "decimal_inch"

        return None, None

    # -------------------------------------------------------------------------
    # INTERPOLATION PRIMITIVES
    # -------------------------------------------------------------------------

    def _interp1d(self, x, pts):
        """
        Linear interpolation (or extrapolation) on sorted (x, y) pairs.
        Returns (value, basis_string, is_extrapolated).
        """
        if len(pts) == 0:
            return 0.0, "No reference points", True

        if len(pts) == 1:
            return pts[0][1], f"Single reference point (x={pts[0][0]:.3f})", False

        # Clamp or extrapolate
        if x <= pts[0][0]:
            if x < pts[0][0]:
                slope = (pts[1][1] - pts[0][1]) / (pts[1][0] - pts[0][0]) if pts[1][0] != pts[0][0] else 0
                val   = pts[0][1] + slope * (x - pts[0][0])
                return max(val, 0.0), f"Extrapolated below min ({pts[0][0]:.3f})", True
            return pts[0][1], f"Exact match at x={pts[0][0]:.3f}", False

        if x >= pts[-1][0]:
            if x > pts[-1][0]:
                slope = (pts[-1][1] - pts[-2][1]) / (pts[-1][0] - pts[-2][0]) if pts[-1][0] != pts[-2][0] else 0
                val   = pts[-1][1] + slope * (x - pts[-1][0])
                return max(val, 0.0), f"Extrapolated above max ({pts[-1][0]:.3f})", True
            return pts[-1][1], f"Exact match at x={pts[-1][0]:.3f}", False

        # Interior interpolation
        for i in range(len(pts) - 1):
            x0, y0 = pts[i]
            x1, y1 = pts[i + 1]
            if x0 <= x <= x1:
                t   = (x - x0) / (x1 - x0) if x1 != x0 else 0
                val = y0 + t * (y1 - y0)
                return val, f"Interpolated between {x0:.3f} and {x1:.3f}", False

        return pts[-1][1], "Fallback to last point", False

    def _interp2d(self, x, y, pts):
        """
        Bilinear-like interpolation on scattered (x, y, energy) points.
        Strategy: group by x-coordinate, interpolate y within each group,
        then interpolate x between groups.
        Returns (value, basis_string, is_extrapolated).
        """
        # Group by x
        x_vals = sorted(set(p[0] for p in pts))
        if len(x_vals) == 1:
            # Only one x level — interpolate y only
            level_pts = sorted([(p[1], p[2]) for p in pts if p[0] == x_vals[0]
                                  and p[1] is not None], key=lambda t: t[0])
            val, basis, extrap = self._interp1d(y, level_pts) if level_pts else (0.0, "no points", True)
            return val, f"1D y-interp at x={x_vals[0]:.3f}: {basis}", extrap

        # Interpolate y at nearest lower and upper x levels
        x_lower = max((xv for xv in x_vals if xv <= x), default=x_vals[0])
        x_upper = min((xv for xv in x_vals if xv >= x), default=x_vals[-1])

        def interp_at_x(xv):
            level_pts = sorted(
                [(p[1], p[2]) for p in pts if p[0] == xv and p[1] is not None],
                key=lambda t: t[0]
            )
            if not level_pts:
                return 0.0, "no points at level", True
            if y is None:
                return level_pts[0][1], "no depth, first row", False
            return self._interp1d(y, level_pts)

        v_lower, b_lower, e_lower = interp_at_x(x_lower)
        v_upper, b_upper, e_upper = interp_at_x(x_upper)

        if x_lower == x_upper:
            extrap = x < x_vals[0] or x > x_vals[-1]
            return v_lower, f"2D: x={x_lower:.3f} | {b_lower}", extrap

        # Linear interpolation between levels
        t   = (x - x_lower) / (x_upper - x_lower)
        val = v_lower + t * (v_upper - v_lower)
        extrap = (x < x_vals[0]) or (x > x_vals[-1])
        basis = (f"2D bilinear: x in [{x_lower:.3f}, {x_upper:.3f}], "
                 f"y={y:.3f}" if y is not None else "2D bilinear: y unknown")
        return max(val, 0.0), basis, extrap

    # -------------------------------------------------------------------------
    # HELPERS
    # -------------------------------------------------------------------------

    def _get_val(self, feature, key):
        val = feature.get(key)
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    def _num(self, val):
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    def _missing(self, reason):
        return {
            "energy_wh":        0.0,
            "estimation_basis": f"Missing data: {reason}",
            "warnings":         [reason],
        }
